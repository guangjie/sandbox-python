#! /Users/jayden/anaconda2/envs/py36/bin/python
import json
import inquirer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

DEV_MODE = False


def main():
    data = load_config_files()
    confirm_send = False
    answers = {}
    common_headers = [
        'Name',
        'Department Shortcut Dimension',
        'Request Type',
        'Project Shortcut',
        'LOTT Type',
        'Detail Type for LOTT',
        'Description',
        'Start Date',
        'Start Time',
        'To Date',
        'To Time',
        'Quantity (Days)'
    ]

    while not confirm_send:
        questions = set_questions(data)
        answers = ask_questions(questions)
        confirm_answer = ask_confirm(answers, questions)
        confirm_send = confirm_answer['confirm']

    set_worksheet(common_headers, answers, data)
    send_email(common_headers, answers, data)


def load_config_files():
    # Loads settings from config file
    with open('config.json') as json_data_file:
        data = json.load(json_data_file)
    return data


def set_questions(data):
    # Data can be assigned directly to cells
    defaults = data["forms"]["leave"]["default"]
    questions = [
        inquirer.Text(
            'name',
            message="What is your Name?",
            default=defaults["name"]
        ),
        inquirer.List(
            'department',
            message="Which Department?",
            choices=['Product', 'Production', 'Sales&Marketing', 'Management', 'Delivery', 'Development'],
            default=defaults["department"]
        ),
        inquirer.List(
            'request_type',
            message="Which Request Type?",
            choices=['Team', 'Project', 'Presales'],
            default=defaults["request_type"]
        ),
        inquirer.List(
            'project_type',
            message="Which Project Shortcut?",
            choices=data["user"]["projects"],
            default=defaults["request_type"]
        ),
        inquirer.List(
            'lott_type',
            message="Which LOTT Type?",
            choices=['Leave', 'OT', 'Travel'],
            default=defaults["lott_type"]
        ),
        inquirer.List(
            'lott_detail_type',
            message="Which Detail Type for LOTT?",
            choices=[
                'OT-Compensation',
                'Travel-Within China',
                'Travel-Outside China',
                'S.Annual Leave',
                'C.Annual Leave',
                'Compensation Leave',
                'Marriage',
                'Sick Leave',
                'Paid Sick Leave',
                'Maternity Leave',
                'Paternity Leave',
                'Personal Leave'
            ],
            default=defaults["lott_detail_type"]
        ),
        inquirer.Text(
            'description',
            message="What is the purpose of the leave?",
            default=defaults["description"]
        ),
        inquirer.Text(
            'start_date',
            message="What is the Start Date? [DD/MM/YYYY]",
        ),
        inquirer.List(
            'start_time',
            message="What is the Start Time?",
            choices=[
                '09.00 AM',
                '12.00 PM'
            ],
            default=defaults["start_time"]
        ),
        inquirer.Text(
            'end_date',
            message="What is the End Date? [DD/MM/YYYY]",
        ),
        inquirer.List(
            'end_time',
            message="What is the End Time?",
            choices=[
                '12.00 PM',
                '06.00 PM'
            ],
            default=defaults["end_time"]
        ),
        inquirer.Text(
            'weekend_ph',
            message="Number of Weekend/PH between Start and End Date?",
            default='0'
        )
    ]

    return questions


def ask_questions(questions):
    answers = inquirer.prompt(questions)
    answers['quantity'] = str(compute_leave_days_applied(
        answers['start_date'],
        answers['end_date'],
        answers['start_time'],
        answers['end_time'],
        answers['weekend_ph'])
    )
    return answers


def ask_confirm(answers, questions):

    print('\n')
    print('--------------------------')
    print('    Leave Application     ')
    print('--------------------------')

    for question in questions:
        print(question.message + ' : ' + answers[question.name])

    print('Total Number of Days : ' + answers['quantity'])
    print('\n')

    question = [
        inquirer.Confirm(
            'confirm',
            message="Proceed with Application?",
            default=True
        )
    ]
    is_confirm = inquirer.prompt(question)
    return is_confirm


def set_worksheet(common_headers, answers, data):
    print('Preparing Excel Spreadsheet...')
    wb = Workbook()
    ws = wb.active
    ws.merge_cells('A1:L1')

    ws['A1'] = data["forms"]["leave"]["title"]
    a1 = ws['A1']

    # Rows can also be appended
    ws.append(common_headers)

    ws.append([
        answers['name'],
        answers['department'],
        answers['request_type'],
        answers['project_type'],
        answers['lott_type'],
        answers['lott_detail_type'],
        answers['description'],
        answers['start_date'],
        answers['start_time'],
        answers['end_date'],
        answers['end_time'],
        answers['quantity']
    ])

    style_font_main_header = Font(name='Calibri (Body)', size='24', bold=True)
    style_font_sub_header = Font(name='Calibri (Body)', size='11', bold=True, color='FFFFFFFF')
    style_fill_light_grey = PatternFill(fill_type='solid', start_color='FFB7B7B7', end_color='FFB7B7B7')
    style_fill_dark_grey = PatternFill(fill_type='solid', start_color='FF808080', end_color='FF808080')
    style_fill_light_blue = PatternFill(fill_type='solid', start_color='FFDDEBF6', end_color='FFDDEBF6')
    style_fill_dark_blue = PatternFill(fill_type='solid', start_color='FF5E9CD3', end_color='FF5E9CD3')
    style_alignment_center = Alignment(horizontal='centerContinuous', vertical='bottom')
    style_thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    a1.alignment = style_alignment_center
    a1.fill = style_fill_light_grey
    a1.font = style_font_main_header
    a1.border = style_thin_border

    for i in range(1, 13):
        ws.cell(row=2, column=i).fill = style_fill_light_grey
        ws.cell(row=2, column=i).font = style_font_sub_header
        ws.cell(row=2, column=i).alignment = style_alignment_center
        ws.cell(row=2, column=i).border = style_thin_border
        ws.cell(row=3, column=i).fill = style_fill_light_blue
        ws.cell(row=3, column=i).alignment = style_alignment_center
        ws.cell(row=3, column=i).border = style_thin_border

    for i in range(8, 12):
        ws.cell(row=2, column=i).fill = style_fill_dark_grey

    ws.cell(row=2, column=7).fill = style_fill_dark_blue

    for col in ws.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length # Make your adjustments here
        ws.column_dimensions[column].width = adjusted_width

    # Save the file
    wb.save(data["user"]["doc_folder"] + 'sample.xlsx')
    print('Excel Spreadsheet Prepared...')


def send_email(common_headers, answers, data):
    print('Preparing Email...')

    if not DEV_MODE:
        from_addr = data["user"]["email"]
        to_addr = data["forms"]["leave"]["emailTo"]
        cc_addr = data["forms"]["leave"]["emailCc"]
    else:
        from_addr = data["dev"]["from_addr"]
        to_addr = data["dev"]["to_addr"]
        cc_addr = data["dev"]["cc_addr"]

    msg = MIMEMultipart()

    msg['From'] = from_addr
    msg['To'] = to_addr
    msg['Cc'] = cc_addr

    msg['Subject'] = answers['lott_type'] + ' Application for ' + answers['name'] + ' [' + answers['start_date'] + ' - ' + answers['end_date'] + ']'

    common_answers = [
        answers['name'],
        answers['department'],
        answers['request_type'],
        answers['project_type'],
        answers['lott_type'],
        answers['lott_detail_type'],
        answers['description'],
        answers['start_date'],
        answers['start_time'],
        answers['end_date'],
        answers['end_time'],
        answers['quantity']
    ]

    style_title_cell = ' style="padding: 8px; background: #ddd; font-weight: bold;" '
    style_normal_cell = ' style="padding: 8px;" '

    body = ''
    body += '<p>Hi Team, </p>'
    body += '<p>I would like to apply for "' + answers['lott_type'] + '" for the period of ' + answers['start_date'] + ' - ' + answers['end_date'] + '.</p>'

    body += '<table>'
    body += '<tbody>'

    for idx, header in enumerate(common_headers):
        body += '<tr>'
        body += '   <td ' + style_title_cell + '>' + header + '</td>'
        body += '   <td ' + style_normal_cell + '>' + common_answers[idx] + '</td>'
        body += '</tr>'

    body += '</tbody>'
    body += '</table>'

    body += '<p>Regards<br/>' + data["user"]["name"] + '</p>'

    msg.attach(MIMEText(body, 'html'))

    filename = 'Attendance Request Form of ' + data["user"]["name"] + '.xlsx'
    attachment = open(data["user"]["doc_folder"] + "sample.xlsx", "rb")

    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= %s" % filename)

    msg.attach(part)

    print('Contacting SMTP Server...')
    server = smtplib.SMTP(data["smtp"]["server"], data["smtp"]["server_port"])
    server.starttls()

    print('Logging into SMTP Server...')
    server.login(from_addr, data["smtp"]["password"])
    text = msg.as_string()

    print('Sending Email...')
    combine_to_addr = to_addr.split(',') + cc_addr.split(',')
    server.sendmail(from_addr, combine_to_addr, text)
    server.quit()

    print('Email Sent...')


def compute_leave_days_applied(str_start_date, str_end_date, str_start_time, str_end_time, str_weekend_ph):
    start_date = get_date(str_start_date)
    end_date = get_date(str_end_date)
    leave_days = end_date - start_date
    num_of_days = float(leave_days.days) + 1

    if str_start_time == '12.00 PM':
        num_of_days = num_of_days - 0.5
    if str_end_time == '12.00 PM':
        num_of_days = num_of_days - 0.5
    if float(str_weekend_ph) != 0.0:
        num_of_days = num_of_days - float(str_weekend_ph)
    return num_of_days


def get_date(str_date):
    list_date = str_date.split('/')
    date_proper = datetime.date(int(list_date[2]), int(list_date[1]), int(list_date[0]))
    return date_proper


main()
